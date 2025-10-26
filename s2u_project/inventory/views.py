import json
import logging
import re
import threading
import time
from decimal import Decimal
from difflib import SequenceMatcher
from typing import Iterable, List, Tuple, Optional, Dict
from uuid import UUID

import requests
from django.core.management import call_command
from django.db import OperationalError, transaction
from django.db.models import Q, Prefetch, Count
from django.http import JsonResponse, HttpResponseNotModified
from django.shortcuts import render
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required, user_passes_test
from django_ratelimit.decorators import ratelimit

from .korona import fetch_product_stocks

logger = logging.getLogger(__name__)
from .models import Product, ProductStock, Store
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth import logout
from .redis_client import r as redis_client
from .redis_client import get_json as redis_get_json, set_json as redis_set_json, setnx as redis_setnx, delete as redis_delete, exists as redis_exists
import threading
import uuid

# Key helpers for global refresh progress
def _refresh_job_key(job_id: str) -> str:
    return f"refresh_job:{job_id}"

def _refresh_lock_key() -> str:
    return "refresh_job:lock"


# Staff check helper used by admin-only views
def _staff_required(user):
    """Return True if the user is authenticated and is staff."""
    return bool(user and user.is_authenticated and user.is_staff)


def _normalize(text: str) -> str:
    """Lowercase and strip non-alphanumeric characters for fuzzy comparisons."""
    return re.sub(r"[^0-9a-z]+", "", text.lower())


def _generate_misspelling_variants(query: str) -> List[str]:
    """Generate common misspelling variants of a query.

    Handles common typos like:
    - Missing/extra letters: tito/titto
    - Phonetic swaps: gray/grey, ph/f
    """
    if not query or len(query) < 3:
        return [query]

    variants = set()
    query_lower = query.lower()

    # Always include original
    variants.add(query_lower)

    # Common phonetic substitutions (only for specific patterns)
    phonetic_map = {
        'gray': 'grey', 'grey': 'gray',
        'ph': 'f', 'ck': 'k',
    }

    for old, new in phonetic_map.items():
        if old in query_lower:
            variants.add(query_lower.replace(old, new))

    # Remove consecutive duplicate letters (titto -> tito)
    dedoubled = re.sub(r'(.)\1+', r'\1', query_lower)
    if dedoubled != query_lower and len(dedoubled) >= 3:
        variants.add(dedoubled)

    return list(variants)


def _score_similarity(query: str, products: Iterable[Product]) -> List[Tuple[Product, float]]:
    """Score products by similarity to query, with lower threshold for misspellings."""
    normalized_query = _normalize(query)
    if not normalized_query:
        return []

    scored: List[Tuple[Product, float]] = []
    for product in products:
        normalized_name = _normalize(product.name)

        # Full name similarity
        score = SequenceMatcher(None, normalized_query, normalized_name).ratio()

        # Also check similarity against each word in the product name
        # This helps with misspellings of specific words
        words = re.split(r'[\s\-]+', normalized_name)
        max_word_score = 0.0
        for word in words:
            if len(word) >= 3:  # Skip very short words
                word_score = SequenceMatcher(None, normalized_query, word).ratio()
                max_word_score = max(max_word_score, word_score)

        # Use the better of the two scores
        final_score = max(score, max_word_score)

        # Lower threshold to 0.50 to catch more misspellings
        if final_score >= 0.50:
            scored.append((product, final_score))

    scored.sort(key=lambda item: item[1], reverse=True)
    return scored[:10]


def _search_products(query: str) -> Tuple[List[Product], List[Product]]:
    """Return ranked product matches and suggestions for a free‑text query.

    Applies multiple strategies: icontains, regex to bridge punctuation, token
    matching, simple phonetic/misspelling variants, and a similarity score for
    tie‑breaking.

    Args:
        query: Raw user input.

    Returns:
        A pair ``(results, suggestions)`` where both are lists of ``Product``.

    Example:
        >>> matches, suggestions = _search_products("titos")  # doctest: +SKIP
        >>> isinstance(matches, list) and isinstance(suggestions, list)
        True
    """
    results: List[Product] = []
    suggestions: List[Product] = []

    # Start with basic filter - exact query
    filters = (
        Q(name__icontains=query)
        | Q(barcode__icontains=query)
        | Q(supplier_name__icontains=query)
    )

    # KEY FIX: Use regex to match query letters with optional special chars between them
    # This allows "titos" to match "TITO'S", "grey" to match "grey", etc.
    # Sanitize the query to get just alphanumeric characters
    sanitized = re.sub(r"[^a-zA-Z0-9\s]", "", query)

    if sanitized and sanitized != query:
        # Simple approach: search for the sanitized version directly
        # Product "TITO'S" in DB, when we search icontains for "TITOS", we need a regex
        # Build a regex pattern: T[^a-zA-Z0-9]*I[^a-zA-Z0-9]*T[^a-zA-Z0-9]*O[^a-zA-Z0-9]*S
        # This matches the letters with any non-alphanumeric chars (including apostrophes) between them
        pattern_chars = [re.escape(c) for c in sanitized.lower()]
        # Allow any non-alphanumeric characters between letters (including apostrophes, spaces, dashes)
        regex_pattern = "[^a-zA-Z0-9]*".join(pattern_chars)
        filters |= Q(name__iregex=regex_pattern) | Q(supplier_name__iregex=regex_pattern)

    # Also try with spaces removed for compound words like "jackdaniels" matching "Jack Daniels"
    no_spaces = sanitized.replace(" ", "")
    if no_spaces and no_spaces != sanitized and len(no_spaces) >= 3:
        pattern_chars = [re.escape(c) for c in no_spaces.lower()]
        regex_pattern = "[^a-zA-Z0-9]*".join(pattern_chars)
        filters |= Q(name__iregex=regex_pattern)

    # Generate misspelling variants (phonetic substitutions, double letters)
    variants = _generate_misspelling_variants(query)
    for variant in variants:
        if variant != query.lower() and len(variant) >= 3:
            filters |= Q(name__icontains=variant)
            # Also add regex version for variants
            variant_sanitized = re.sub(r"[^a-zA-Z0-9]", "", variant)
            if variant_sanitized and len(variant_sanitized) >= 3:
                pattern_chars = [re.escape(c) for c in variant_sanitized.lower()]
                regex_pattern = "[^a-zA-Z0-9]*".join(pattern_chars)
                filters |= Q(name__iregex=regex_pattern)

    # Split into tokens and search each word independently
    # This helps with multi-word queries like "titos vodka"
    tokens = [t.strip() for t in re.split(r"[\s\-]+", query) if len(t.strip()) >= 2]
    for token in tokens:
        # Add original token
        filters |= Q(name__icontains=token)

        # Add regex pattern for token (handles special chars)
        token_clean = re.sub(r"[^a-zA-Z0-9]", "", token)
        if token_clean and len(token_clean) >= 2:
            pattern_chars = [re.escape(c) for c in token_clean.lower()]
            regex_pattern = "[^a-zA-Z0-9]*".join(pattern_chars)
            filters |= Q(name__iregex=regex_pattern) | Q(supplier_name__iregex=regex_pattern)

        # Add misspelling variants of the token
        if len(token_clean) >= 3:
            token_variants = _generate_misspelling_variants(token_clean)
            for tv in token_variants:
                if tv != token_clean and len(tv) >= 3:
                    filters |= Q(name__icontains=tv)
                    # Add regex pattern for variant
                    pattern_chars = [re.escape(c) for c in tv]
                    regex_pattern = "[^a-zA-Z0-9]*".join(pattern_chars)
                    filters |= Q(name__iregex=regex_pattern)

    exact_q = Q(barcode__iexact=query) | Q(barcodes__code__iexact=query)
    if query.isdigit():
        exact_q |= Q(number=int(query))
    normalized_query = _normalize(query)

    candidates = (
        Product.objects.filter(filters | exact_q | Q(barcodes__code__icontains=query))
        .distinct()
        .only("number", "name", "barcode", "supplier_name")
    )

    ranked: List[Tuple[int, float, Product]] = []
    for product in candidates:
        priority = 0
        normalized_name = _normalize(product.name)
        lowers_name = product.name.lower()
        query_lower = query.lower()

        # Exact matches get highest priority
        if product.barcode and product.barcode == query:
            priority += 10
        if query.isdigit() and product.number == int(query):
            priority += 9

        # Normalized exact match (handles "titos" matching "TITO'S")
        if normalized_name == normalized_query:
            priority += 8

        # Word boundary match at start (whole word match)
        if normalized_name.startswith(normalized_query):
            priority += 7

        # Check if query matches start of any word in product name
        words = re.split(r'[\s\-]+', lowers_name)
        for word in words:
            word_normalized = _normalize(word)
            if word_normalized.startswith(normalized_query):
                priority += 6
                break

        # Regular starts-with match
        if lowers_name.startswith(query_lower):
            priority += 5

        # Contains match
        if query_lower in lowers_name:
            priority += 3

        # Token matching - each word in query appears somewhere in product name
        query_tokens = [_normalize(t) for t in re.split(r'\s+', query) if len(t) > 1]
        if query_tokens:
            matches = sum(1 for token in query_tokens if token in normalized_name)
            if matches == len(query_tokens):
                priority += 4  # All tokens match
            elif matches > 0:
                priority += 2  # Partial token match

        # Supplier name bonus
        if product.supplier_name and query_lower in product.supplier_name.lower():
            priority += 1

        # Use similarity ratio for tie-breaking
        ratio = SequenceMatcher(None, normalized_query, normalized_name).ratio()
        ranked.append((priority, ratio, product))

    ranked.sort(key=lambda item: (-item[0], -item[1], item[2].name))
    results = [product for _, _, product in ranked]

    if len(results) < 10:
        # Use iterator to avoid loading all products into memory
        candidate_qs = (
            Product.objects.exclude(pk__in=[p.pk for p in results])
            .only("number", "name", "barcode", "supplier_name")
            .order_by("name")[:500]  # Limit to first 500 for similarity check
        )
        scored = _score_similarity(query, candidate_qs)
        suggestions = [product for product, _ in scored if product not in results]

    return results, suggestions


@login_required
def product_search(request):
    """Render the interactive product search page.

    Displays search results and suggestions along with active stores so the
    user can quickly fetch stock data.

    Example:
        GET /inventory/?q=titos
    """
    query = request.GET.get("q", "").strip()
    stores = Store.objects.filter(active=True).order_by("name")
    store_data = list(stores.values("id", "name", "number"))
    store_data_json = json.dumps(store_data)

    results: List[Product] = []
    suggestions: List[Product] = []

    if query:
        results, suggestions = _search_products(query)

    return render(
        request,
        "inventory/product_search.html",
        {
            "query": query,
            "results": results,
            "suggestions": suggestions,
            "stores": stores,
            "store_data": store_data,
            "store_data_json": store_data_json,
            "active_tab": "inventory",
        },
    )


def home(request):
    """Dashboard landing page.

    Adds a grouped view of weekly order lists per store, with lists
    sorted by target date (newest first) and an Edit link per list.
    """
    # If not authenticated, render public landing page with login CTA
    if not request.user.is_authenticated:
        return render(
            request,
            "inventory/landing.html",
            {
                "active_tab": "home",
            },
        )
    from .models import WeeklyOrderList

    # Filters from query params
    store_param = (request.GET.get("store") or "").strip()
    date_from_raw = (request.GET.get("date_from") or "").strip()
    date_to_raw = (request.GET.get("date_to") or "").strip()

    from datetime import date

    def parse_date(value: str):
        try:
            return date.fromisoformat(value)
        except Exception:
            return None

    date_from = parse_date(date_from_raw)
    date_to = parse_date(date_to_raw)

    # Base queryset for weekly lists, filtered by date if provided
    wl_base = WeeklyOrderList.objects.all()
    if date_from:
        wl_base = wl_base.filter(target_date__gte=date_from)
    if date_to:
        wl_base = wl_base.filter(target_date__lte=date_to)
    wl_base = wl_base.order_by("-target_date", "-created_at").annotate(item_count=Count("items"))

    # Stores to show
    stores_all = Store.objects.filter(active=True).order_by("name")

    selected_store = None
    if store_param and store_param.lower() != "all":
        try:
            selected_store = stores_all.get(pk=int(store_param))
            stores_all = stores_all.filter(pk=selected_store.pk)
        except (ValueError, Store.DoesNotExist):
            # Fallback: try by store number
            selected_store = Store.objects.filter(number=str(store_param), active=True).first()
            if selected_store:
                stores_all = stores_all.filter(pk=selected_store.pk)

    # Prefetch weekly lists for each (filtered) store
    store_qs = stores_all.prefetch_related(Prefetch("weekly_lists", queryset=wl_base))

    return render(
        request,
        "inventory/home.html",
        {
            "active_tab": "home",
            "stores": store_qs,
            "stores_all": Store.objects.filter(active=True).order_by("name"),
            "selected_store": selected_store,
            "date_from": date_from_raw,
            "date_to": date_to_raw,
            "can_refresh": bool(request.user.is_staff),
        },
    )


@ratelimit(key='user_or_ip', rate='100/m', method='GET', block=True)
@login_required
def product_search_api(request):
    """Search products API with rate limiting: 100 requests per minute per user."""
    query = request.GET.get("q", "").strip()
    store_id = request.GET.get("store")
    if not query:
        return JsonResponse({"matches": [], "suggestions": []})

    matches, suggestions = _search_products(query)
    store = None
    if store_id:
        try:
            store = Store.objects.get(pk=store_id)
        except (Store.DoesNotExist, ValueError):
            store = Store.objects.filter(number=str(store_id)).first()

    def serialize(product: Product) -> dict:
        return {
            "number": product.number,
            "name": product.name,
            "barcode": product.barcode,
            "supplier_name": product.supplier_name,
        }

    payload = {
        "matches": [serialize(product) for product in matches],
        "suggestions": [serialize(product) for product in suggestions],
    }
    if store:
        payload["store"] = {
            "id": store.pk,
            "name": store.name,
            "number": store.number,
        }

    return JsonResponse(payload)


def about(request):
    """Public About page with brief info and login CTA."""
    return render(request, "inventory/about.html", {"active_tab": "about"})


@ratelimit(key='user_or_ip', rate='50/m', method='GET', block=True)
@login_required
@require_GET
def product_stock_api(request):
    """Return current stock for a product across stores.

    Query params:
        product: Product number (required)
        store: Optional store ID or store number to limit the scope
        force: Set to '1' to bypass cache (admins bypass automatically)

    Example:
        # All stores
        GET /api/stock/?product=123

        # Single store by ID
        GET /api/stock/?product=123&store=4
    """
    product_number = request.GET.get("product")
    store_identifier = (request.GET.get("store") or "").strip()

    if not product_number:
        return JsonResponse({"error": "Parameter 'product' is required."}, status=400)

    try:
        product = Product.objects.get(number=int(product_number))
    except (Product.DoesNotExist, ValueError):
        return JsonResponse({"error": "Product not found."}, status=404)

    if not product.korona_id:
        return JsonResponse(
            {"error": "Product is missing Korona integration data."}, status=400
        )

    store_scope = "all" if not store_identifier else "single"
    target_store = None
    if store_scope == "single":
        try:
            target_store = Store.objects.get(pk=store_identifier)
        except (Store.DoesNotExist, ValueError):
            target_store = Store.objects.filter(number=str(store_identifier)).first()
        if not target_store:
            return JsonResponse({"error": "Store not found."}, status=404)
        if not target_store.korona_id:
            return JsonResponse(
                {"error": "Store is missing Korona integration data."}, status=400
            )

    # Admins get fresh data, regular users get cached data
    # Also allow force refresh via ?force=1 parameter
    force_refresh = request.user.is_staff or request.GET.get('force') == '1'
    api_failed = False
    try:
        payload = fetch_product_stocks(product.korona_id, force_refresh=force_refresh)
    except requests.RequestException as exc:
        # API failed - fall back to cached DB data
        logger.warning(f"Korona API failed for product {product.number}, using cached data: {exc}")
        api_failed = True
        payload = None

    active_stores = list(Store.objects.filter(active=True))
    store_map = {store.korona_id: store for store in active_stores}
    stock_entries: List[dict] = []
    seen_store_ids: set[int] = set()

    def update_entry(store_obj: Store, defaults: dict) -> None:
        # Retry logic with exponential backoff for database locks
        for attempt in range(3):
            try:
                with transaction.atomic():
                    stock, _ = ProductStock.objects.update_or_create(
                        product=product,
                        store=store_obj,
                        defaults=defaults,
                    )
                break
            except OperationalError as exc:
                if attempt == 2:
                    raise exc
                # Exponential backoff: 50ms, 100ms
                time.sleep(0.05 * (2 ** attempt))
        seen_store_ids.add(store_obj.pk)
        stock_entries.append(
            {
                "store": {
                    "id": store_obj.pk,
                    "name": store_obj.name,
                    "number": store_obj.number,
                },
                "stock": {
                    "actual": float(stock.actual),
                    "lent": float(stock.lent),
                    "ordered": float(stock.ordered),
                    "max_level": float(stock.max_level),
                    "reorder_level": float(stock.reorder_level),
                    "average_purchase_price": float(stock.average_purchase_price),
                    "listed": stock.listed,
                    "updated_at": stock.updated_at.isoformat(),
                },
            }
        )

    zero_defaults = {
        "actual": Decimal("0"),
        "lent": Decimal("0"),
        "max_level": Decimal("0"),
        "ordered": Decimal("0"),
        "reorder_level": Decimal("0"),
        "average_purchase_price": Decimal("0"),
        "listed": False,
    }

    results = (payload or {}).get("results") or [] if not api_failed else []

    # If API failed and no results, try to get cached data from DB
    if api_failed or not results:
        logger.info(f"Using cached stock data for product {product.number}")
        cached_stocks = ProductStock.objects.filter(product=product).select_related('store')

        for stock_obj in cached_stocks:
            if stock_obj.store not in active_stores:
                continue

            stock_entries.append({
                "store": {
                    "id": stock_obj.store.pk,
                    "name": stock_obj.store.name,
                    "number": stock_obj.store.number,
                },
                "stock": {
                    "actual": float(stock_obj.actual),
                    "lent": float(stock_obj.lent),
                    "ordered": float(stock_obj.ordered),
                    "max_level": float(stock_obj.max_level),
                    "reorder_level": float(stock_obj.reorder_level),
                    "average_purchase_price": float(stock_obj.average_purchase_price),
                    "listed": stock_obj.listed,
                    "updated_at": stock_obj.updated_at.isoformat(),
                },
                "cached": True,
            })
            seen_store_ids.add(stock_obj.store.pk)

    # Process API results if available
    for entry in results:
        warehouse = entry.get("warehouse") or {}
        warehouse_id = warehouse.get("id")
        if not warehouse_id:
            continue
        try:
            korona_store_id = UUID(str(warehouse_id))
        except ValueError:
            continue

        store_obj = store_map.get(korona_store_id)
        if not store_obj:
            continue
        if store_scope == "single" and target_store and store_obj.pk != target_store.pk:
            continue

        amount = entry.get("amount") or {}
        defaults = {
            "actual": Decimal(str(amount.get("actual", "0") or "0")),
            "lent": Decimal(str(amount.get("lent", "0") or "0")),
            "max_level": Decimal(str(amount.get("maxLevel", "0") or "0")),
            "ordered": Decimal(str(amount.get("ordered", "0") or "0")),
            "reorder_level": Decimal(str(amount.get("reorderLevel", "0") or "0")),
            "average_purchase_price": Decimal(
                str(entry.get("averagePurchasePrice", "0") or "0")
            ),
            "listed": bool(entry.get("listed", False)),
        }
        update_entry(store_obj, defaults)

    if store_scope == "single" and target_store:
        if target_store.pk not in seen_store_ids:
            update_entry(target_store, zero_defaults)
        stock_entries.sort(key=lambda entry: (entry["store"].get("name") or "").lower())

        return JsonResponse(
            {
                "product": {"number": product.number, "name": product.name},
                "stocks": stock_entries,
                "cached": api_failed,
            }
        )

    for store_obj in active_stores:
        if store_obj.pk not in seen_store_ids:
            update_entry(store_obj, zero_defaults)

    ProductStock.objects.filter(product=product).exclude(store__pk__in=seen_store_ids).delete()
    stock_entries.sort(key=lambda entry: (entry["store"].get("name") or "").lower())

    # Build barcodes list (primary + additional)
    try:
        extra_codes = list(product.barcodes.values_list('code', flat=True))  # type: ignore[attr-defined]
    except Exception:
        extra_codes = []
    barcodes = []
    if product.barcode:
        barcodes.append(product.barcode)
    for code in extra_codes:
        if code and code not in barcodes:
            barcodes.append(code)

    return JsonResponse(
        {
            "product": {"number": product.number, "name": product.name},
            "stocks": stock_entries,
            "cached": api_failed,
            "barcodes": barcodes,
        }
    )


@ratelimit(key='user_or_ip', rate='60/m', method='GET', block=True)
@login_required
@require_GET
def product_refresh_api(request):
    """Trigger background sync of stores or products and return counts.

    Query params:
        sync: 'stores' or 'products' (default: 'products')

    Example:
        GET /api/products/refresh/?sync=stores
    """
    sync_type = request.GET.get('sync', 'products')
    async_mode = request.GET.get('async') == '1'

    try:
        if sync_type == 'stores':
            # Sync organizational units (stores)
            if async_mode:
                threading.Thread(target=lambda: call_command("sync_stores"), daemon=True).start()
                total = Store.objects.filter(active=True).count()
                return JsonResponse({"ok": True, "queued": True, "total": total, "type": "stores"}, status=202)
            call_command("sync_stores")
            total = Store.objects.filter(active=True).count()
            return JsonResponse({"ok": True, "total": total, "type": "stores"})
        else:
            # Default: sync products
            # Prevent concurrent/too‑frequent syncs
            lock_key = "product_refresh:lock"
            if async_mode:
                try:
                    got = redis_setnx(lock_key, "1", ex=900)
                except Exception:
                    got = True
                if got:
                    def _worker():
                        try:
                            call_command("load_products", skip_csv=True)
                        finally:
                            try:
                                redis_delete(lock_key)
                            except Exception:
                                pass
                    threading.Thread(target=_worker, daemon=True).start()
                total = Product.objects.count()
                return JsonResponse({"ok": True, "queued": True, "total": total, "type": "products"}, status=202)
            else:
                call_command("load_products", skip_csv=True)
                total = Product.objects.count()
                return JsonResponse({"ok": True, "total": total, "type": "products"})
    except Exception as exc:  # pylint: disable=broad-except
        return JsonResponse(
            {"ok": False, "error": f"Failed to refresh {sync_type}: {exc}"}, status=500
        )


# ===== Global async refresh (stores, products, stocks, monthly sales) =====

def _update_progress(job_id: str, payload: Dict) -> None:
    key = _refresh_job_key(job_id)
    current = redis_get_json(key, {}) or {}
    current.update(payload)
    redis_set_json(key, current, ex=3600)  # keep for up to 1 hour


def _run_refresh_job(job_id: str, user_id: int) -> None:
    """Background worker to refresh stores and products only (lightweight)."""
    try:
        # Acquire simple lock to avoid concurrent global refreshes
        # Acquire lock with NX
        redis_setnx(_refresh_lock_key(), job_id, ex=1800)

        _update_progress(job_id, {"step": "init", "message": "Starting refresh...", "done": False})

        # Step 1: Sync stores
        _update_progress(job_id, {"step": "stores", "message": "Syncing stores...", "progress": 5})
        call_command("sync_stores")
        stores_qs = Store.objects.filter(active=True)
        stores_count = stores_qs.count()
        _update_progress(job_id, {"stores": stores_count, "progress": 40, "message": f"Stores synced: {stores_count}"})

        # Step 2: Sync products
        _update_progress(job_id, {"step": "products", "message": "Syncing products...", "progress": 45})
        call_command("load_products", skip_csv=True)
        products_qs = Product.objects.exclude(korona_id__isnull=True)
        products_count = products_qs.count()
        _update_progress(job_id, {"products": products_count, "progress": 90, "message": f"Products synced: {products_count}"})

        # Done
        _update_progress(job_id, {
            "step": "done",
            "message": "Refresh complete.",
            "progress": 100,
            "done": True,
        })
    except Exception as exc:  # pylint: disable=broad-except
        _update_progress(job_id, {
            "step": "error",
            "message": f"Failed: {exc}",
            "error": str(exc),
            "done": True,
        })
    finally:
        # Release lock if we own it
        try:
            if (redis_client.get(_refresh_lock_key()) or "") == job_id:
                redis_delete(_refresh_lock_key())
        except Exception:
            pass


@login_required
@user_passes_test(_staff_required)
def refresh_all_start_api(request):
    """Start an async refresh job. Returns a job ID for polling."""
    # Prevent parallel runs
    if redis_exists(_refresh_lock_key()):
        return JsonResponse({"ok": False, "error": "A refresh is already running. Please wait."}, status=409)

    job_id = uuid.uuid4().hex
    # Seed job status
    redis_set_json(_refresh_job_key(job_id), {"step": "queued", "message": "Queued...", "progress": 0, "done": False}, ex=3600)

    # Spawn background thread
    t = threading.Thread(target=_run_refresh_job, args=(job_id, request.user.id), daemon=True)
    t.start()

    return JsonResponse({"ok": True, "job": job_id})


@login_required
@user_passes_test(_staff_required)
def refresh_all_status_api(request):
    """Return status of a refresh job given ?job=<job_id>."""
    job_id = (request.GET.get("job") or "").strip()
    if not job_id:
        return JsonResponse({"ok": False, "error": "Missing job id"}, status=400)
    data = redis_get_json(_refresh_job_key(job_id))
    if not data:
        return JsonResponse({"ok": False, "error": "Job not found"}, status=404)
    return JsonResponse({"ok": True, **data})


@ratelimit(key='user_or_ip', rate='900/m', method='GET', block=True)
@login_required
@require_GET
def monthly_sales_api(request):
    """
    Fetch monthly sales for a product at specific stores with caching.
    OPTIMIZED: Fetches receipts once for all stores instead of per-store.

    Query params:
        product: product number (required)
        stores: comma-separated store IDs (optional, defaults to all active)
        force: set to '1' to bypass cache (optional)
    """
    from .models import MonthlySales
    from datetime import timedelta

    product_number = request.GET.get("product", "").strip()
    store_ids_param = request.GET.get("stores", "").strip()
    force_refresh = request.GET.get("force") == "1"

    logger.info(f"[MONTHLY SALES API] Request for product={product_number}, stores={store_ids_param}, force={force_refresh}")

    if not product_number:
        logger.warning("[MONTHLY SALES API] Missing product parameter")
        return JsonResponse({"error": "Parameter 'product' is required."}, status=400)

    try:
        product = Product.objects.get(number=int(product_number))
        logger.info(f"[MONTHLY SALES API] Found product: {product.number} - {product.name}")
    except (Product.DoesNotExist, ValueError):
        logger.error(f"[MONTHLY SALES API] Product {product_number} not found")
        return JsonResponse({"error": "Product not found."}, status=404)

    if not product.korona_id:
        logger.error(f"[MONTHLY SALES API] Product {product.number} missing Korona ID")
        return JsonResponse({"error": "Product missing Korona integration."}, status=400)

    # Determine target stores
    if store_ids_param:
        store_ids = [int(sid.strip()) for sid in store_ids_param.split(",") if sid.strip().isdigit()]
        stores = Store.objects.filter(pk__in=store_ids, active=True)
        logger.info(f"[MONTHLY SALES API] Filtering to {len(store_ids)} specific stores: {store_ids}")
    else:
        stores = Store.objects.filter(active=True)
        logger.info(f"[MONTHLY SALES API] Using all active stores")

    # Filter stores that have Korona integration
    stores_with_korona = [s for s in stores if s.korona_id]
    logger.info(f"[MONTHLY SALES API] Found {len(stores_with_korona)} stores with Korona integration")

    if not stores_with_korona:
        logger.warning(f"[MONTHLY SALES API] No stores with Korona integration found")
        return JsonResponse({
            "product": {"number": product.number, "name": product.name},
            "sales": {},
        })

    from .redis_client import r as redis_client

    sales_data = {}
    stores_needing_calculation = []

    # Two-tier caching: Redis (fast) -> Database (persistent) -> API (slow)
    if not force_refresh:
        logger.info(f"[MONTHLY SALES API] Checking cache for {len(stores_with_korona)} stores...")
        for store in stores_with_korona:
            # Layer 1: Check Redis cache (fastest)
            redis_key = f"monthly_sales:{product.number}:{store.id}"
            cached_val = redis_client.get(redis_key)
            cached_qty = int(cached_val) if (cached_val is not None and str(cached_val).isdigit()) else None

            if cached_qty is not None:
                logger.info(f"[MONTHLY SALES API] ✓ Redis HIT for product {product.number} at store {store.number}: {cached_qty}")
                sales_data[store.id] = cached_qty
                continue

            # Layer 2: Check database cache
            try:
                cached_sale = MonthlySales.objects.get(product=product, store=store)
                # Check if stale (> 30 minutes)
                if not cached_sale.is_stale:
                    logger.info(f"[MONTHLY SALES API] ✓ DB HIT for product {product.number} at store {store.number}: {cached_sale.quantity_sold} (age: {cached_sale.calculated_at})")
                    # Update Redis cache for next time (30 min TTL = 1800 seconds)
                    redis_client.set(redis_key, int(cached_sale.quantity_sold), ex=3600)
                    sales_data[store.id] = cached_sale.quantity_sold
                else:
                    logger.info(f"[MONTHLY SALES API] ✗ DB STALE for product {product.number} at store {store.number} (age: {cached_sale.calculated_at})")
                    stores_needing_calculation.append(store)
            except MonthlySales.DoesNotExist:
                logger.info(f"[MONTHLY SALES API] ✗ NO CACHE for product {product.number} at store {store.number}")
                stores_needing_calculation.append(store)
    else:
        logger.info(f"[MONTHLY SALES API] Force refresh enabled - bypassing all cache")
        stores_needing_calculation = stores_with_korona

    # If we need to calculate for any stores, fetch receipts ONCE and process for all stores
    if stores_needing_calculation:
        logger.info(f"[MONTHLY SALES API] Need to calculate for {len(stores_needing_calculation)} stores")
        try:
            # Fetch sales for ALL stores at once (much faster!)
            from .korona import calculate_monthly_sales_bulk
            logger.info(f"[MONTHLY SALES API] Starting bulk API fetch for product {product.number}...")

            bulk_sales = calculate_monthly_sales_bulk(
                str(product.korona_id),
                [(s.id, str(s.korona_id)) for s in stores_needing_calculation],
                days=30
            )

            logger.info(f"[MONTHLY SALES API] ✓ Bulk fetch complete. Results: {bulk_sales}")

            # Update both Redis and database cache, plus sales_data
            for store in stores_needing_calculation:
                qty = bulk_sales.get(store.id, 0)

                # Update Redis cache (30 min TTL = 1800 seconds)
                redis_key = f"monthly_sales:{product.number}:{store.id}"
                redis_client.set(redis_key, int(qty), ex=3600)
                logger.info(f"[MONTHLY SALES API] ✓ Saved to Redis: product {product.number} at store {store.number} = {qty}")

                # Update database cache (persistent)
                MonthlySales.objects.update_or_create(
                    product=product,
                    store=store,
                    defaults={
                        "quantity_sold": qty,
                        "days_calculated": 30,
                    }
                )
                logger.info(f"[MONTHLY SALES API] ✓ Saved to DB: product {product.number} at store {store.number} = {qty}")

                sales_data[store.id] = qty

        except Exception as exc:
            logger.error(f"[MONTHLY SALES API] ✗ FAILED to calculate bulk monthly sales for product {product.number}: {exc}", exc_info=True)
            # Try to use stale cache for stores that failed
            for store in stores_needing_calculation:
                try:
                    cached_sale = MonthlySales.objects.get(product=product, store=store)
                    logger.warning(f"[MONTHLY SALES API] Using stale cache for store {store.number}: {cached_sale.quantity_sold}")
                    sales_data[store.id] = cached_sale.quantity_sold
                except MonthlySales.DoesNotExist:
                    logger.warning(f"[MONTHLY SALES API] No cache available for store {store.number}")
                    pass

    logger.info(f"[MONTHLY SALES API] ✓ Response ready: {len(sales_data)} stores with data")

    # Build payload and ETag for HTTP caching
    import json as _json
    import hashlib as _hashlib
    payload = {"product": {"number": product.number, "name": product.name}, "sales": sales_data}
    payload_str = _json.dumps(payload, sort_keys=True, separators=(",", ":"))
    etag = 'W/"' + _hashlib.md5(payload_str.encode("utf-8")).hexdigest() + '"'

    # If client sent matching ETag and not force refresh, return 304
    inm = request.META.get("HTTP_IF_NONE_MATCH")
    if not force_refresh and inm and inm.strip() == etag:
        resp = HttpResponseNotModified()
        resp["ETag"] = etag
        resp["Cache-Control"] = "private, max-age=60, stale-while-revalidate=120"
        return resp

    resp = JsonResponse(payload)
    resp["ETag"] = etag
    if force_refresh:
        resp["Cache-Control"] = "no-store"
    else:
        resp["Cache-Control"] = "private, max-age=60, stale-while-revalidate=120"
    return resp


@ratelimit(key='user_or_ip', rate='120/m', method='GET', block=True)
@login_required
@require_GET
def monthly_sales_bulk_api(request):
    """Return monthly sales for multiple products in one call.

    Query params:
      - products: comma-separated product numbers (required)
      - stores: comma-separated store IDs (optional, defaults to all active)
      - force: '1' to bypass cache (optional)

    Response:
      {
        "sales": { "<product_number>": { "<store_id>": qty, ... }, ... },
        "missing": [<product_number>...]
      }
    """
    from .models import MonthlySales
    import json as _json
    import hashlib as _hashlib

    products_param = (request.GET.get("products") or "").strip()
    if not products_param:
        return JsonResponse({"error": "Parameter 'products' is required."}, status=400)

    # Parse and de-duplicate product numbers
    try:
        product_numbers = list({int(p.strip()) for p in products_param.split(',') if p.strip()})
    except ValueError:
        return JsonResponse({"error": "Invalid 'products' list."}, status=400)

    # Hard cap per request to keep latency reasonable
    if len(product_numbers) > 500:
        return JsonResponse({"error": "Too many products; max 500 per request."}, status=400)

    stores_param = (request.GET.get("stores") or "").strip()
    force_refresh = request.GET.get("force") == "1"

    # Resolve stores
    if stores_param:
        try:
            store_ids = [int(s.strip()) for s in stores_param.split(',') if s.strip()]
        except ValueError:
            return JsonResponse({"error": "Invalid 'stores' list."}, status=400)
        stores = Store.objects.filter(pk__in=store_ids, active=True)
    else:
        stores = Store.objects.filter(active=True)

    stores_with_korona = [s for s in stores if s.korona_id]
    if not stores_with_korona:
        return JsonResponse({"sales": {}, "missing": product_numbers})

    # Fetch Products
    products = {p.number: p for p in Product.objects.filter(number__in=product_numbers)}
    missing = [n for n in product_numbers if n not in products]

    from .redis_client import r as redis_client
    sales_out: dict[int, dict[int, int]] = {}

    # First pass: try Redis and DB cache
    stores_needing: dict[int, list[Store]] = {}

    from datetime import timedelta
    fresh_cutoff = timezone.now() - timedelta(minutes=30)

    for num, product in products.items():
        if not product.korona_id:
            missing.append(num)
            continue
        sales_out[num] = {}
        if not force_refresh:
            for st in stores_with_korona:
                # Redis
                redis_key = f"monthly_sales:{num}:{st.id}"
                cached_val = redis_client.get(redis_key)
                if cached_val is not None and str(cached_val).isdigit():
                    sales_out[num][st.id] = int(cached_val)
                    continue
                # DB cache
                try:
                    ms = MonthlySales.objects.only("quantity_sold", "calculated_at").get(product=product, store=st)
                    if ms.calculated_at >= fresh_cutoff:
                        sales_out[num][st.id] = int(ms.quantity_sold)
                        redis_client.set(redis_key, int(ms.quantity_sold), ex=3600)
                        continue
                except MonthlySales.DoesNotExist:
                    pass
                # needs calc
                stores_needing.setdefault(num, []).append(st)
        else:
            stores_needing[num] = stores_with_korona.copy()

    # Second pass: calculate missing/stale using Korona in a per-product bulk call
    from .korona import calculate_monthly_sales_bulk
    for num, need_list in stores_needing.items():
        product = products.get(num)
        if not product or not need_list:
            continue
        try:
            bulk_sales = calculate_monthly_sales_bulk(
                str(product.korona_id),
                [(s.id, str(s.korona_id)) for s in need_list],
                days=30,
            )
            # Persist to Redis + DB
            for st in need_list:
                qty = int(bulk_sales.get(st.id, 0))
                sales_out.setdefault(num, {})[st.id] = qty
                redis_client.set(f"monthly_sales:{num}:{st.id}", qty, ex=3600)
                MonthlySales.objects.update_or_create(
                    product=product,
                    store=st,
                    defaults={"quantity_sold": qty, "days_calculated": 30},
                )
        except Exception as exc:  # noqa: BLE001
            logger.error("[MONTHLY BULK] failed for product %s: %s", num, exc, exc_info=True)
            # Fall back to stale DB values when present
            try:
                stale = MonthlySales.objects.filter(product=product, store_id__in=[s.id for s in need_list])
                for ms in stale:
                    sales_out.setdefault(num, {})[int(ms.store_id)] = int(ms.quantity_sold)
            except Exception:
                pass

    # Build payload + ETag
    payload = {"sales": sales_out, "missing": missing}
    payload_str = _json.dumps(payload, sort_keys=True, separators=(",", ":"))
    etag = 'W/"' + _hashlib.md5(payload_str.encode("utf-8")).hexdigest() + '"'

    inm = request.META.get("HTTP_IF_NONE_MATCH")
    if not force_refresh and inm and inm.strip() == etag:
        resp = HttpResponseNotModified()
        resp["ETag"] = etag
        resp["Cache-Control"] = "private, max-age=60, stale-while-revalidate=120"
        return resp

    resp = JsonResponse(payload)
    resp["ETag"] = etag
    resp["Cache-Control"] = "no-store" if force_refresh else "private, max-age=60, stale-while-revalidate=120"
    return resp


@login_required
def weekly_list_create(request):
    """View to create a new weekly order list."""
    from datetime import date

    if request.method == "POST":
        store_id = request.POST.get("store")
        target_date = request.POST.get("target_date")

        if not store_id or not target_date:
            return render(
                request,
                "inventory/weekly_list_create.html",
                {
                    "stores": Store.objects.filter(active=True).order_by("name"),
                    "default_date": date.today().isoformat(),
                    "error": "Please select both store and date.",
                    "active_tab": "weekly",
                },
            )

        try:
            store = Store.objects.get(pk=store_id)
        except Store.DoesNotExist:
            return render(
                request,
                "inventory/weekly_list_create.html",
                {
                    "stores": Store.objects.filter(active=True).order_by("name"),
                    "default_date": date.today().isoformat(),
                    "error": "Invalid store selected.",
                    "active_tab": "weekly",
                },
            )

        from django.shortcuts import redirect
        from .models import WeeklyOrderList

        order_list = WeeklyOrderList.objects.create(
            store=store,
            target_date=target_date,
        )

        return redirect("inventory:weekly_list_detail", list_id=order_list.id)

    # GET request
    from datetime import date
    stores = Store.objects.filter(active=True).order_by("name")
    return render(
        request,
        "inventory/weekly_list_create.html",
        {
            "stores": stores,
            "default_date": date.today().isoformat(),
            "active_tab": "weekly",
        },
    )


@login_required
def weekly_list_detail(request, list_id):
    """View to display and manage a weekly order list."""
    from django.shortcuts import get_object_or_404
    from .models import WeeklyOrderList

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)

    is_admin = bool(request.user.is_staff)
    can_edit = (order_list.finalized_at is None) or is_admin
    show_admin_cols = is_admin or bool(order_list.finalized_at)

    # For finalized lists, show transfer items first
    if order_list.finalized_at:
        # Sort: transfers first (with transfer_from set), then by product name
        from django.db.models import Case, When, Value, IntegerField
        items = order_list.items.select_related("product", "transfer_from").prefetch_related("product__barcodes").annotate(
            has_transfer=Case(
                When(transfer_from__isnull=False, transfer_bottles__gt=0, then=Value(0)),
                default=Value(1),
                output_field=IntegerField()
            )
        ).order_by('has_transfer', 'product__name')
    else:
        items = order_list.items.select_related("product", "transfer_from").prefetch_related("product__barcodes").all()

    # Only load other store stocks for admins
    other_stores_qs = Store.objects.none()
    other_store_ids = []
    stock_map: dict[tuple[int, int], float] = {}

    if is_admin:
        # Prepare other active stores (exclude current) for cross-store stock columns
        other_stores_qs = Store.objects.filter(active=True).exclude(pk=order_list.store_id).order_by("number")
        other_store_ids = list(other_stores_qs.values_list("id", flat=True))

        # Build a stock map: (product_id, store_id) -> actual using bulk fetch
        product_ids = [item.product_id for item in items]
        if product_ids and other_store_ids:
            # Bulk fetch all stocks in one query
            stocks = ProductStock.objects.filter(
                product_id__in=product_ids,
                store_id__in=other_store_ids
            ).values_list("product_id", "store_id", "actual")
            stock_map = {(pid, sid): float(actual) for pid, sid, actual in stocks}

    # --- Server-side prefill of monthly sales to reduce initial flicker ---
    current_sales_map: dict[int, int] = {}
    other_sales_map: dict[tuple[int, int], int] = {}
    try:
        from .models import MonthlySales
        from django.utils import timezone
        from datetime import timedelta
        product_ids = [item.product_id for item in items]
        sales_qs = MonthlySales.objects.filter(product_id__in=product_ids)
        target_store_ids = [order_list.store_id]
        if is_admin:
            target_store_ids += other_store_ids
        sales_qs = sales_qs.filter(store_id__in=target_store_ids)
        thirty_min_ago = timezone.now() - timedelta(minutes=30)
        sales_qs = sales_qs.filter(calculated_at__gte=thirty_min_ago)
        for s in sales_qs.only("product_id", "store_id", "quantity_sold", "calculated_at"):
            if s.store_id == order_list.store_id:
                current_sales_map[int(s.product_id)] = int(s.quantity_sold)
            else:
                other_sales_map[(int(s.product_id), int(s.store_id))] = int(s.quantity_sold)
    except Exception as _exc:
        logger.warning("[weekly_detail] Monthly prefill skipped: %s", _exc)

    # Serialize items for JavaScript
    items_data = [
        [
            item.product.number,
            {
                "id": item.id,
                "product_number": item.product.number,
                "product_name": item.product.name,
                "barcode": item.product.barcode,
                "barcodes": [b.code for b in getattr(item.product, 'barcodes').all()] if hasattr(item.product, 'barcodes') else ([item.product.barcode] if item.product.barcode else []),
                "supplier_name": item.product.supplier_name,
                "on_shelf": item.on_shelf,
                "monthly_needed": item.monthly_needed,
                "system_stock": float(item.system_stock),
                "transfer_from_id": item.transfer_from_id,
                "transfer_from_number": (item.transfer_from.number if item.transfer_from else None),
                "transfer_bottles": item.transfer_bottles,
                "joe": item.joe,
                "bt": item.bt,
                "sqw": item.sqw,
                "has_transfer": bool(item.transfer_from_id and item.transfer_bottles and item.transfer_bottles > 0),
                "other_stocks": {sid: stock_map.get((item.product_id, sid), 0.0) for sid in other_store_ids} if is_admin else {},
                "monthly_sales": int(current_sales_map.get(item.product_id, 0)),
                "other_monthly_sales": (
                    {sid: other_sales_map.get((item.product_id, sid), 0) for sid in other_store_ids}
                    if is_admin else {}
                ),
            },
        ]
        for item in items
    ]

    # stores for transfer dropdown (exclude current store)
    stores = list(
        Store.objects.filter(active=True).exclude(pk=order_list.store_id).order_by("name").values("id", "name", "number")
    )

    # Supplier list for filters
    supplier_set = set()
    for it in items:
        supplier_set.add(it.product.supplier_name or "—")
    suppliers = sorted(supplier_set, key=lambda s: (s == "—", s.lower()))

    return render(
        request,
        "inventory/weekly_list_detail.html",
        {
            "order_list": order_list,
            "items_json": json.dumps(items_data),
            "active_tab": "weekly",
            "is_admin": is_admin,
            "show_admin_cols": show_admin_cols,
            "stores_json": json.dumps(stores),
            "other_stores": other_stores_qs,
            "other_stores_json": json.dumps(list(other_stores_qs.values("id", "number", "name"))),
            "suppliers_json": json.dumps(suppliers),
            "can_edit": can_edit,
        },
    )


@login_required
def weekly_search_api(request, list_id):
    """API endpoint to search products for weekly list (store-specific)."""
    from django.shortcuts import get_object_or_404
    from .models import WeeklyOrderList

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)
    query = request.GET.get("q", "").strip()

    if not query:
        return JsonResponse({"matches": [], "suggestions": []})

    matches, suggestions = _search_products(query)

    def serialize(product: Product) -> dict:
        return {
            "number": product.number,
            "name": product.name,
            "barcode": product.barcode,
            "supplier_name": product.supplier_name,
        }

    return JsonResponse(
        {
            "matches": [serialize(product) for product in matches],
            "suggestions": [serialize(product) for product in suggestions],
        }
    )


@login_required
def weekly_add_item_api(request, list_id):
    """API endpoint to add a product to a weekly list."""
    import json as json_lib
    from django.shortcuts import get_object_or_404
    from .models import WeeklyOrderList, WeeklyOrderItem

    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)
    if order_list.finalized_at and not request.user.is_staff:
        return JsonResponse({"error": "This list has been finalized."}, status=403)

    try:
        data = json_lib.loads(request.body)
        product_number = data.get("product_number")

        if not product_number:
            return JsonResponse({"error": "Product number is required"}, status=400)

        try:
            product = Product.objects.get(number=int(product_number))
        except (Product.DoesNotExist, ValueError):
            return JsonResponse({"error": "Product not found"}, status=404)

        # Check if item already exists
        existing_item = WeeklyOrderItem.objects.filter(
            order_list=order_list, product=product
        ).first()

        if existing_item:
            # Increment on_shelf count
            existing_item.on_shelf += 1
            existing_item.save()
            item = existing_item
        else:
            # Fetch system stock for this store (prefer fresh; fallback to DB cached)
            system_stock = Decimal("0")
            if product.korona_id and order_list.store.korona_id:
                try:
                    payload = fetch_product_stocks(product.korona_id, force_refresh=True)
                    results = (payload or {}).get("results") or []

                    for entry in results:
                        warehouse = entry.get("warehouse") or {}
                        warehouse_id = warehouse.get("id")
                        if warehouse_id:
                            try:
                                korona_store_id = UUID(str(warehouse_id))
                                if korona_store_id == order_list.store.korona_id:
                                    amount = entry.get("amount") or {}
                                    system_stock = Decimal(str(amount.get("actual", "0") or "0"))
                                    break
                            except ValueError:
                                continue
                except requests.RequestException:
                    # Network/API failure: use last known DB value if available and schedule a retry
                    cached_ps = ProductStock.objects.filter(product=product, store=order_list.store).only("actual").first()
                    if cached_ps:
                        system_stock = cached_ps.actual
                    retry_in = 45
                    def _retry_worker(prod_id: int, store_id: int, item_pk: int, order_list_pk: int):
                        try:
                            from .models import Product as ProdModel, Store as StoreModel, WeeklyOrderItem
                            prod = ProdModel.objects.get(pk=prod_id)
                            store = StoreModel.objects.get(pk=store_id)
                            payload2 = fetch_product_stocks(prod.korona_id, force_refresh=True)
                            results2 = (payload2 or {}).get("results") or []
                            for entry in results2:
                                warehouse = entry.get("warehouse") or {}
                                wid = warehouse.get("id")
                                if not wid:
                                    continue
                                try:
                                    if UUID(str(wid)) != store.korona_id:
                                        continue
                                except ValueError:
                                    continue
                                amount = entry.get("amount") or {}
                                new_stock = Decimal(str(amount.get("actual", "0") or "0"))
                                # Persist: ProductStock and the weekly item row
                                with transaction.atomic():
                                    ProductStock.objects.update_or_create(
                                        product=prod,
                                        store=store,
                                        defaults={
                                            "actual": new_stock,
                                            "lent": Decimal("0"),
                                            "max_level": Decimal("0"),
                                            "ordered": Decimal("0"),
                                            "reorder_level": Decimal("0"),
                                            "average_purchase_price": Decimal("0"),
                                            "listed": True,
                                        },
                                    )
                                    try:
                                        witem = WeeklyOrderItem.objects.get(pk=item_pk, order_list_id=order_list_pk)
                                        witem.system_stock = new_stock
                                        witem.save(update_fields=["system_stock"])
                                    except WeeklyOrderItem.DoesNotExist:
                                        pass
                                break
                        except Exception as exc:  # noqa: BLE001
                            logger.warning("[retry] system stock retry failed for product=%s store=%s: %s", prod_id, store_id, exc)


            # Create new item
            item = WeeklyOrderItem.objects.create(
                order_list=order_list,
                product=product,
                on_shelf=1,
                monthly_needed=0,
                system_stock=system_stock,
            )
            # If a retry was scheduled, start it now with the real item id
            try:
                if 'retry_in' in locals():
                    t = threading.Timer(retry_in, _retry_worker, args=(product.pk, order_list.store.pk, item.pk, order_list.pk))
                    t.daemon = True
                    t.start()
            except Exception as exc:
                logger.warning("[retry] scheduling failed: %s", exc)

        # Only build cross-store data for admins
        other_map = {}

        if request.user.is_staff:
            # Build cross-store stocks for this product across other active stores (exclude current)
            other_stores_qs = Store.objects.filter(active=True).exclude(pk=order_list.store_id)
            other_store_ids = list(other_stores_qs.values_list("id", flat=True))
            other_map = {int(sid): 0.0 for sid in other_store_ids}
            if other_store_ids:
                for ps in ProductStock.objects.filter(product=product, store_id__in=other_store_ids).only("store_id", "actual"):
                    other_map[int(ps.store_id)] = float(ps.actual)

        return JsonResponse(
            {
                "id": item.id,
                "product_number": product.number,
                "product_name": product.name,
                "barcode": product.barcode,
                "supplier_name": product.supplier_name,
                "on_shelf": item.on_shelf,
                "monthly_needed": item.monthly_needed,
                "system_stock": float(item.system_stock),
                "transfer_from_id": item.transfer_from_id,
                "transfer_from_number": (item.transfer_from.number if item.transfer_from else None),
                "transfer_bottles": item.transfer_bottles,
                "joe": item.joe,
                "bt": item.bt,
                "sqw": item.sqw,
                "other_stocks": other_map,
                **({"retry_scheduled": True, "retry_in_seconds": retry_in} if 'retry_in' in locals() else {}),
            }
        )

    except json_lib.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)


@login_required
def weekly_update_item_api(request, list_id, item_id):
    """API endpoint to update a weekly list item."""
    import json as json_lib
    from django.shortcuts import get_object_or_404
    from .models import WeeklyOrderList, WeeklyOrderItem

    if request.method != "PATCH":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)
    if order_list.finalized_at and not request.user.is_staff:
        return JsonResponse({"error": "This list has been finalized."}, status=403)
    # Accept either WeeklyOrderItem.id or Product.number in the URL for robustness
    try:
        item = WeeklyOrderItem.objects.get(pk=item_id, order_list=order_list)
    except WeeklyOrderItem.DoesNotExist:
        # Fallback: treat item_id as product.number under this list
        try:
            item = WeeklyOrderItem.objects.select_related("product").get(
                order_list=order_list, product__number=item_id
            )
        except WeeklyOrderItem.DoesNotExist:
            return JsonResponse({"error": "Item not found."}, status=404)

    try:
        data = json_lib.loads(request.body)

        if "on_shelf" in data:
            item.on_shelf = max(0, int(data["on_shelf"]))

        if "system_stock" in data:
            item.system_stock = Decimal(str(data["system_stock"]))

        if "monthly_needed" in data:
            item.monthly_needed = max(0, int(data["monthly_needed"]))

        # Admin-only fields
        if request.user.is_staff:
            if "transfer_from" in data:
                store_id_raw = data.get("transfer_from")
                if store_id_raw in (None, "", 0, "0"):
                    item.transfer_from = None
                else:
                    try:
                        st = Store.objects.get(pk=int(store_id_raw))
                        item.transfer_from = st
                    except (Store.DoesNotExist, ValueError):
                        pass
            for key in ("transfer_bottles", "joe", "bt", "sqw"):
                if key in data:
                    try:
                        setattr(item, key, max(0, int(data.get(key) or 0)))
                    except (TypeError, ValueError):
                        pass
        else:
            # If a non-admin tries to update admin-only fields, reject
            if any(k in data for k in ("transfer_from", "transfer_bottles", "joe", "bt", "sqw")):
                return JsonResponse({"error": "Not authorized"}, status=403)

        item.save()

        # Only build cross-store data for admins
        other_map = {}

        if request.user.is_staff:
            # Build cross-store stocks for this product across other active stores (exclude current)
            other_stores_qs = Store.objects.filter(active=True).exclude(pk=order_list.store_id)
            other_store_ids = list(other_stores_qs.values_list("id", flat=True))
            other_map = {int(sid): 0.0 for sid in other_store_ids}
            if other_store_ids:
                for ps in ProductStock.objects.filter(product=item.product, store_id__in=other_store_ids).only("store_id", "actual"):
                    other_map[int(ps.store_id)] = float(ps.actual)

        return JsonResponse(
            {
                "id": item.id,
                "product_number": item.product.number,
                "product_name": item.product.name,
                "barcode": item.product.barcode,
                "supplier_name": item.product.supplier_name,
                "on_shelf": item.on_shelf,
                "monthly_needed": item.monthly_needed,
                "system_stock": float(item.system_stock),
                "transfer_from_id": item.transfer_from_id,
                "transfer_from_number": (item.transfer_from.number if item.transfer_from else None),
                "transfer_bottles": item.transfer_bottles,
                "joe": item.joe,
                "bt": item.bt,
                "sqw": item.sqw,
                "other_stocks": other_map,
            }
        )

    except (json_lib.JSONDecodeError, ValueError) as exc:
        return JsonResponse({"error": str(exc)}, status=400)


@login_required
def weekly_delete_item_api(request, list_id, item_id):
    """API endpoint to delete a weekly list item."""
    from django.shortcuts import get_object_or_404
    from .models import WeeklyOrderList, WeeklyOrderItem

    if request.method != "DELETE":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)
    if order_list.finalized_at and not request.user.is_staff:
        return JsonResponse({"error": "This list has been finalized."}, status=403)
    item = get_object_or_404(WeeklyOrderItem, pk=item_id, order_list=order_list)

    item.delete()

    return JsonResponse({"ok": True})


@login_required
def weekly_export_excel(request, list_id):
    """Export weekly list to Excel format."""
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse
    from .models import WeeklyOrderList
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)
    # Use iterator for memory efficiency on large lists
    items = order_list.items.select_related("product", "transfer_from").order_by("product__name").iterator(chunk_size=500)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Order List"

    # Build title + metadata rows
    from django.utils import timezone as dj_tz
    user_name = getattr(request.user, "get_full_name", lambda: "")() or getattr(request.user, "username", "")
    title_text = f"Weekly Order List - {order_list.store.name} (#{order_list.store.number})"
    subtitle_text = f"Week of {order_list.target_date} • Generated on {dj_tz.now().astimezone().strftime('%Y-%m-%d %H:%M')} by {user_name}"

    max_cols = 12
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_cols)
    ws.cell(row=1, column=1, value=title_text).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=subtitle_text).font = Font(color="666666")

    # Add headers
    headers = [
        "Product #",
        "Product Name",
        "Barcode",
        "Supplier",
        "System Stock",
        "On Shelf",
        "Weekly Needed",
        "Transfer From",
        "Transfer Bottles",
        "Joe",
        "BT",
        "SQW",
    ]
    ws.append([None] * len(headers))  # placeholder for row 3 (merged rows took 1-2)
    ws.append(headers)  # row 4

    # Style header row
    header_row_idx = 4
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze panes below header
    ws.freeze_panes = ws["A5"]

    # Add data rows
    from openpyxl.styles import Border, Side
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

    for idx, item in enumerate(items, start=1):
        row = [
            item.product.number,
            item.product.name,
            item.product.barcode or "",
            item.product.supplier_name or "",
            float(item.system_stock),
            item.on_shelf,
            item.monthly_needed,
            (item.transfer_from.number if item.transfer_from else ""),
            item.transfer_bottles,
            item.joe,
            item.bt,
            item.sqw,
        ]
        ws.append(row)
        r = header_row_idx + idx
        if idx % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = zebra_fill
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            # Alignment and number formats
            if c in (1, 5, 6, 7, 9, 10, 11, 12):
                cell.alignment = Alignment(horizontal="center")
            if c == 5:
                cell.number_format = "0.00"

    # Auto-adjust column widths with caps
    from openpyxl.utils import get_column_letter
    widths = {1: 11, 2: 28, 3: 14, 4: 22, 5: 12, 6: 11, 7: 13, 8: 12, 9: 15, 10: 8, 11: 8, 12: 8}
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 14)

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"weekly_list_{order_list.store.name.replace(' ', '_')}_{order_list.target_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def weekly_export_excel_custom(request, list_id):
    """Export a filtered weekly list to Excel (admins only).

    Accepts POST body or form-encoded ``payload`` JSON with keys like
    ``columns``, ``supplier``, ``has`` and ``other_stores``.

    Example:
        curl -X POST \
             -H 'Content-Type: application/json' \
             -d '{"columns":["product_number","product_name"],"has":["joe"]}' \
             http://localhost:8000/weekly/1/export/excel/custom/
    """
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse
    from .models import WeeklyOrderList
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    import json as json_lib

    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    # Only admins can use custom filtered export
    if not request.user.is_staff:
        return JsonResponse({"error": "Filtered export is only available for administrators"}, status=403)

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)
    payload_raw = request.POST.get("payload") or request.body
    try:
        payload = json_lib.loads(payload_raw)
    except Exception:
        payload = {}

    # Columns and filters
    sel_cols = payload.get("columns") or []
    supplier_filter = payload.get("supplier") or None
    has_fields = payload.get("has") or []
    other_store_ids = payload.get("other_stores") or []

    # Enforce permissions: employees cannot export admin-only columns unless finalized
    admin_cols = {"transfer_from", "transfer_bottles", "joe", "bt", "sqw"}
    is_admin = bool(request.user.is_staff)
    if not is_admin and not order_list.finalized_at:
        sel_cols = [c for c in sel_cols if c not in admin_cols]

    items_qs = order_list.items.select_related("product", "transfer_from")
    if supplier_filter:
        if supplier_filter == "—":
            items_qs = items_qs.filter(product__supplier_name__in=["", None])
        else:
            items_qs = items_qs.filter(product__supplier_name=supplier_filter)
    for field in has_fields:
        if field in admin_cols or field in {"transfer_bottles", "joe", "bt", "sqw"}:
            items_qs = items_qs.filter(**{f"{field}__gt": 0})
    items_qs = items_qs.order_by("product__name")

    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Weekly List"

    # Header mapping
    col_labels = {
        "product_number": "Product #",
        "product_name": "Product Name",
        "barcode": "Barcode",
        "supplier": "Supplier",
        "system_stock": "System Stock",
        "on_shelf": "On Shelf",
        "monthly_needed": "Monthly Needed",
        "transfer_from": "Transfer From",
        "transfer_bottles": "Transfer Bottles",
        "joe": "Joe",
        "bt": "BT",
        "sqw": "SQW",
    }

    # Title + metadata
    from django.utils import timezone as dj_tz
    user_name = getattr(request.user, "get_full_name", lambda: "")() or getattr(request.user, "username", "")
    title_text = f"Weekly List (Filtered) - {order_list.store.name} (#{order_list.store.number})"
    subtitle_bits = [f"Week of {order_list.target_date}"]
    if supplier_filter:
        subtitle_bits.append(f"Supplier: {supplier_filter}")
    if has_fields:
        subtitle_bits.append("Has: " + ", ".join(has_fields))
    subtitle_bits.append(f"Generated on {dj_tz.now().astimezone().strftime('%Y-%m-%d %H:%M')} by {user_name}")
    subtitle_text = " • ".join(subtitle_bits)

    # Build headers
    headers = [col_labels[c] for c in sel_cols if c in col_labels]
    other_qs = Store.objects.filter(pk__in=other_store_ids).order_by("number")
    other_cols = [(st.pk, st.number) for st in other_qs]
    headers.extend([str(num) for _, num in other_cols])

    max_cols = max(1, len(headers))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_cols)
    ws.cell(row=1, column=1, value=title_text).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=subtitle_text).font = Font(color="666666")

    ws.append([None] * max_cols)  # placeholder row 3
    ws.append(headers)  # header row (4)

    # Style header + freeze panes
    header_row_idx = 4
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws["A5"]

    # Preload other store stocks to reduce queries
    product_ids = list(items_qs.values_list("product_id", flat=True))
    stock_map = {}
    if product_ids and other_cols:
        for ps in ProductStock.objects.filter(product_id__in=product_ids, store_id__in=[sid for sid, _ in other_cols]).only("product_id", "store_id", "actual"):
            stock_map[(ps.product_id, ps.store_id)] = float(ps.actual)

    from openpyxl.styles import Border, Side
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

    for idx, it in enumerate(items_qs, start=1):
        row = []
        for key in sel_cols:
            if key == "product_number":
                row.append(it.product.number)
            elif key == "product_name":
                row.append(it.product.name)
            elif key == "barcode":
                row.append(it.product.barcode or "")
            elif key == "supplier":
                row.append(it.product.supplier_name or "")
            elif key == "system_stock":
                row.append(float(it.system_stock))
            elif key == "on_shelf":
                row.append(it.on_shelf)
            elif key == "monthly_needed":
                row.append(it.monthly_needed)
            elif key == "transfer_from":
                row.append(it.transfer_from.number if it.transfer_from else "")
            elif key == "transfer_bottles":
                row.append(it.transfer_bottles)
            elif key == "joe":
                row.append(it.joe)
            elif key == "bt":
                row.append(it.bt)
            elif key == "sqw":
                row.append(it.sqw)
        # other store stocks
        for sid, _ in other_cols:
            row.append(stock_map.get((it.product_id, sid), 0.0))
        ws.append(row)
        r = header_row_idx + idx
        if idx % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = zebra_fill
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            # center numeric-ish columns
            cell.alignment = Alignment(horizontal="center")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"weekly_list_filtered_{order_list.store.number}_{order_list.target_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def weekly_export_pdf(request, list_id):
    """Export weekly list to PDF format - ADMIN ONLY."""
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse
    from .models import WeeklyOrderList
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO

    # Only admins can export to PDF
    if not request.user.is_staff:
        return JsonResponse({"error": "PDF export is only available for administrators"}, status=403)

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)
    # Use iterator for memory efficiency on large lists
    items = order_list.items.select_related("product", "transfer_from").order_by("product__name").iterator(chunk_size=500)

    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5 * inch)

    # Container for PDF elements
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#2563EB"),
        spaceAfter=12,
        alignment=1,  # Center
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Normal"],
        fontSize=12,
        textColor=colors.gray,
        spaceAfter=20,
        alignment=1,  # Center
    )

    # Add title and subtitle
    title = Paragraph(f"Weekly Order List - {order_list.store.name}", title_style)
    # Add generated line and store name (already in title)
    from django.utils import timezone as dj_tz
    user_name = getattr(request.user, "get_full_name", lambda: "")() or getattr(request.user, "username", "")
    subtitle = Paragraph(
        f"Week of {order_list.target_date.strftime('%B %d, %Y')} • Generated on {dj_tz.now().astimezone().strftime('%Y-%m-%d %H:%M')} by {user_name}",
        subtitle_style,
    )
    elements.append(title)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.2 * inch))

    # Prepare table data (use Paragraph for wrapping)
    headers = [
        "Product #",
        "Product Name",
        "Barcode",
        "Supplier",
        "System Stock",
        "On Shelf",
        "Weekly Needed",
        "Transfer From",
        "Transfer Bottles",
        "Joe",
        "BT",
        "SQW",
    ]

    table_rows = []
    cell_style = ParagraphStyle(
        "cell",
        parent=styles["Normal"],
        fontSize=8.5,
        leading=11,
        wordWrap="CJK",
    )

    def P(text):
        return Paragraph(str(text) if text not in (None, "") else "—", cell_style)

    for item in items:
        table_rows.append(
            [
                P(item.product.number),
                P(item.product.name),
                P(item.product.barcode or "—"),
                P(item.product.supplier_name or "—"),
                P(f"{float(item.system_stock):.2f}"),
                P(item.on_shelf),
                P(item.monthly_needed),
                P(item.transfer_from.number if item.transfer_from else "—"),
                P(item.transfer_bottles),
                P(item.joe),
                P(item.bt),
                P(item.sqw),
            ]
        )

    # Build dynamic column widths that fit within page width
    base_widths = [0.9, 2.8, 1.4, 2.6, 1.0, 0.9, 1.1, 1.1, 1.1, 0.8, 0.8, 0.8]  # inches
    col_widths = [w * inch for w in base_widths[: len(headers)]]
    table = Table([headers] + table_rows, colWidths=col_widths, repeatRows=1)

    # Style table
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9.5),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 1), (-1, -1), 8.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                ("ALIGN", (0, 1), (0, -1), "CENTER"),  # Product #
                ("ALIGN", (4, 1), (6, -1), "CENTER"),  # numeric stocks
                ("ALIGN", (7, 1), (-1, -1), "CENTER"),  # admin numeric
            ]
        )
    )

    elements.append(table)

    # Build PDF
    doc.build(elements)

    # Prepare response
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    filename = f"weekly_list_{order_list.store.name.replace(' ', '_')}_{order_list.target_date}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


@login_required
def weekly_export_pdf_custom(request, list_id):
    """Export a filtered weekly list to PDF.

    Accepts a JSON ``payload`` with the same keys used by the custom Excel
    export. Non-admin users can only export non-admin columns unless the list
    is finalized.

    Example:
        curl -X POST -H 'Content-Type: application/json' \
             -d '{"columns":["product_name","on_shelf"]}' \
             http://localhost:8000/weekly/1/export/pdf/custom/
    """
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse
    from .models import WeeklyOrderList
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    import json as json_lib

    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)
    payload_raw = request.POST.get("payload") or request.body
    try:
        payload = json_lib.loads(payload_raw)
    except Exception:
        payload = {}

    sel_cols = payload.get("columns") or []
    supplier_filter = payload.get("supplier") or None
    has_fields = payload.get("has") or []
    other_store_ids = payload.get("other_stores") or []

    admin_cols = {"transfer_from", "transfer_bottles", "joe", "bt", "sqw"}
    is_admin = bool(request.user.is_staff)
    if not is_admin and not order_list.finalized_at:
        sel_cols = [c for c in sel_cols if c not in admin_cols]

    items_qs = order_list.items.select_related("product", "transfer_from")
    if supplier_filter:
        if supplier_filter == "—":
            items_qs = items_qs.filter(product__supplier_name__in=["", None])
        else:
            items_qs = items_qs.filter(product__supplier_name=supplier_filter)
    for field in has_fields:
        if field in admin_cols or field in {"transfer_bottles", "joe", "bt", "sqw"}:
            items_qs = items_qs.filter(**{f"{field}__gt": 0})
    items_qs = items_qs.order_by("product__name")

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.4 * inch)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading2"], textColor=colors.HexColor("#2563EB"))
    subtitle_style = ParagraphStyle("Sub", parent=styles["Normal"], textColor=colors.gray, fontSize=9)
    from django.utils import timezone as dj_tz
    user_name = getattr(request.user, "get_full_name", lambda: "")() or getattr(request.user, "username", "")
    elements.append(Paragraph(f"Weekly List (Filtered) - {order_list.store.name}", title_style))
    elements.append(Paragraph(f"Week of {order_list.target_date.strftime('%B %d, %Y')} • Generated on {dj_tz.now().astimezone().strftime('%Y-%m-%d %H:%M')} by {user_name}", subtitle_style))
    elements.append(Spacer(1, 0.12 * inch))

    col_labels = {
        "product_number": "Product #",
        "product_name": "Product Name",
        "barcode": "Barcode",
        "supplier": "Supplier",
        "system_stock": "System Stock",
        "on_shelf": "On Shelf",
        "monthly_needed": "Monthly Needed",
        "transfer_from": "Transfer From",
        "transfer_bottles": "Transfer Bottles",
        "joe": "Joe",
        "bt": "BT",
        "sqw": "SQW",
    }

    other_qs = Store.objects.filter(pk__in=other_store_ids).order_by("number")
    other_cols = [(st.pk, st.number) for st in other_qs]

    headers = [col_labels[c] for c in sel_cols if c in col_labels]
    headers.extend([str(num) for _, num in other_cols])

    stock_map = {}
    product_ids = list(items_qs.values_list("product_id", flat=True))
    if product_ids and other_cols:
        for ps in ProductStock.objects.filter(product_id__in=product_ids, store_id__in=[sid for sid, _ in other_cols]).only("product_id", "store_id", "actual"):
            stock_map[(ps.product_id, ps.store_id)] = float(ps.actual)

    # Build rows with wrapped text via Paragraphs
    table_rows = []
    cell_style = ParagraphStyle(
        "cell",
        parent=styles["Normal"],
        fontSize=8.5,
        leading=11,
        wordWrap="CJK",
    )
    def P(text):
        return Paragraph(str(text) if text not in (None, "") else "—", cell_style)

    for it in items_qs:
        row = []
        for key in sel_cols:
            if key == "product_number":
                row.append(P(it.product.number))
            elif key == "product_name":
                row.append(P(it.product.name))
            elif key == "barcode":
                row.append(P(it.product.barcode or "—"))
            elif key == "supplier":
                row.append(P(it.product.supplier_name or "—"))
            elif key == "system_stock":
                row.append(P(f"{float(it.system_stock):.2f}"))
            elif key == "on_shelf":
                row.append(P(it.on_shelf))
            elif key == "weekly_needed":
                row.append(P(it.weekly_needed))
            elif key == "transfer_from":
                row.append(P(it.transfer_from.number if it.transfer_from else "—"))
            elif key == "transfer_bottles":
                row.append(P(it.transfer_bottles))
            elif key == "joe":
                row.append(P(it.joe))
            elif key == "bt":
                row.append(P(it.bt))
            elif key == "sqw":
                row.append(P(it.sqw))
        for sid, _ in other_cols:
            row.append(P(f"{stock_map.get((it.product_id, sid), 0.0):.2f}"))
        table_rows.append(row)

    # Column width planning
    # Assign base widths per header; cap to page width
    base_map = {
        "Product #": 0.9,
        "Product Name": 2.8,
        "Barcode": 1.4,
        "Supplier": 2.6,
        "System Stock": 1.0,
        "On Shelf": 1.0,
        "Weekly Needed": 1.1,
        "Transfer From": 1.1,
        "Transfer Bottles": 1.2,
        "Joe": 0.9,
        "BT": 0.9,
        "SQW": 0.9,
    }
    base_widths = [base_map.get(h, 0.9) for h in headers]
    col_widths = [w * inch for w in base_widths]
    table = Table([headers] + table_rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9.5),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 1), (-1, -1), 8.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
            ("ALIGN", (0, 1), (0, -1), "CENTER"),
            ("ALIGN", (4, 1), (-1, -1), "CENTER"),
        ])
    )
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    filename = f"weekly_list_filtered_{order_list.store.number}_{order_list.target_date}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ===== Admin-only: finalize and delete lists =====

@login_required
@user_passes_test(lambda u: u.is_staff)
def weekly_export_custom(request, list_id):
    """
    Custom export for Joe/BT/SQW/Transfer with filtered rows.

    Query params:
        type: 'joe', 'bt', 'sqw', or 'transfer'
        format: 'excel' or 'pdf'
    """
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse
    from .models import WeeklyOrderList

    export_type = request.GET.get('type', '').lower()
    export_format = request.GET.get('format', 'excel').lower()

    logger.info(f"[EXPORT] Custom export request: type={export_type}, format={export_format}, list_id={list_id}")

    if export_type not in ['joe', 'bt', 'sqw', 'transfer']:
        return JsonResponse({"error": "Invalid export type"}, status=400)

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)

    # Filter items based on export type
    if export_type == 'transfer':
        # Transfer: has transfer_from AND transfer_bottles > 0
        items = order_list.items.select_related("product", "transfer_from").filter(
            transfer_from__isnull=False,
            transfer_bottles__gt=0
        ).order_by("product__name")
        columns = ["Product Name", "Barcode", "Supplier", "Transfer From", "Transfer Bottles"]
    else:
        # Joe/BT/SQW: respective field > 0
        filter_kwargs = {f"{export_type}__gt": 0}
        items = order_list.items.select_related("product").filter(**filter_kwargs).order_by("product__name")
        columns = ["Product Name", "Barcode", "Supplier", "System Stock", export_type.upper()]

    logger.info(f"[EXPORT] Found {items.count()} items matching filter")

    if export_format == 'excel':
        return _export_custom_excel(request, order_list, items, export_type, columns)
    else:
        return _export_custom_pdf(request, order_list, items, export_type, columns)


def _export_custom_excel(request, order_list, items, export_type, columns):
    """Generate Excel for custom export."""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.utils import timezone as dj_tz

    wb = Workbook()
    ws = wb.active
    ws.title = f"{export_type.upper()} Export"

    # Title
    user_name = getattr(request.user, "get_full_name", lambda: "")() or getattr(request.user, "username", "")
    title_text = f"{export_type.upper()} Export - {order_list.store.name} (#{order_list.store.number})"
    subtitle_text = f"Week of {order_list.target_date} • Generated on {dj_tz.now().astimezone().strftime('%Y-%m-%d %H:%M')} by {user_name}"

    max_cols = len(columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_cols)
    ws.cell(row=1, column=1, value=title_text).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=subtitle_text).font = Font(color="666666")

    # Headers
    ws.append([None] * len(columns))
    ws.append(columns)

    header_row_idx = 4
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = ws["A5"]

    # Data rows
    for item in items:
        if export_type == 'transfer':
            row_data = [
                item.product.name,
                item.product.barcode or '',
                item.product.supplier_name or '',
                f"{item.transfer_from.name} (#{item.transfer_from.number})" if item.transfer_from else '',
                item.transfer_bottles or 0
            ]
        else:
            row_data = [
                item.product.name,
                item.product.barcode or '',
                item.product.supplier_name or '',
                item.system_stock or 0,
                getattr(item, export_type, 0) or 0
            ]
        ws.append(row_data)

    # Auto-size columns (skip merged cells)
    from openpyxl.cell.cell import MergedCell
    for col_idx, column in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = None
        for cell in column:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            # Get column letter from first non-merged cell
            if column_letter is None:
                column_letter = cell.column_letter
            # Calculate max length
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{export_type}_export_{order_list.store.number}_{order_list.target_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    logger.info(f"[EXPORT] Excel generated successfully: {filename}")
    return response


def _export_custom_pdf(request, order_list, items, export_type, columns):
    """Generate PDF for custom export."""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from django.utils import timezone as dj_tz
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)

    elements = []
    styles = getSampleStyleSheet()

    # Title
    user_name = getattr(request.user, "get_full_name", lambda: "")() or getattr(request.user, "username", "")
    title_text = f"{export_type.upper()} Export - {order_list.store.name} (#{order_list.store.number})"
    subtitle_text = f"Week of {order_list.target_date} • Generated on {dj_tz.now().astimezone().strftime('%Y-%m-%d %H:%M')} by {user_name}"

    title_style = ParagraphStyle(name='CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#2563EB'))
    elements.append(Paragraph(title_text, title_style))
    elements.append(Paragraph(subtitle_text, styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))

    # Table data
    table_data = [columns]

    for item in items:
        if export_type == 'transfer':
            row_data = [
                item.product.name[:40],
                item.product.barcode or '',
                item.product.supplier_name[:20] if item.product.supplier_name else '',
                f"{item.transfer_from.name} (#{item.transfer_from.number})" if item.transfer_from else '',
                str(item.transfer_bottles or 0)
            ]
        else:
            row_data = [
                item.product.name[:40],
                item.product.barcode or '',
                item.product.supplier_name[:20] if item.product.supplier_name else '',
                str(item.system_stock or 0),
                str(getattr(item, export_type, 0) or 0)
            ]
        table_data.append(row_data)

    # Create table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f"{export_type}_export_{order_list.store.number}_{order_list.target_date}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    logger.info(f"[EXPORT] PDF generated successfully: {filename}")
    return response

@login_required
@user_passes_test(_staff_required)
def weekly_finalize_list(request, list_id):
    """Finalize a weekly list (admins only).

    Sets ``finalized_at`` and ``finalized_by``; thereafter employees cannot
    edit the list. Only accepts POST.

    Example:
        POST /weekly/123/finalize/
    """
    from django.shortcuts import get_object_or_404, redirect
    from .models import WeeklyOrderList

    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)
    if not order_list.finalized_at:
        order_list.finalized_at = timezone.now()
        order_list.finalized_by = request.user
        order_list.save(update_fields=["finalized_at", "finalized_by"])
        messages.success(request, "List finalized.")
    return redirect("inventory:weekly_list_detail", list_id=list_id)


@login_required
@user_passes_test(_staff_required)
def weekly_delete_list(request, list_id):
    """Delete a weekly list (admins only). POST only.

    Example:
        POST /weekly/123/delete/
    """
    from django.shortcuts import get_object_or_404, redirect
    from .models import WeeklyOrderList

    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)
    store_name = order_list.store.name
    order_list.delete()
    messages.success(request, f"Deleted list for {store_name}.")
    return redirect("inventory:home")


@login_required
@user_passes_test(_staff_required)
def weekly_unfinalize_list(request, list_id):
    """Unfinalize a weekly list (admins only). POST only.

    Removes ``finalized_at`` and ``finalized_by`` to re‑enable editing.

    Example:
        POST /weekly/123/unfinalize/
    """
    from django.shortcuts import get_object_or_404, redirect
    from .models import WeeklyOrderList

    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    order_list = get_object_or_404(WeeklyOrderList, pk=list_id)
    if order_list.finalized_at:
        order_list.finalized_at = None
        order_list.finalized_by = None
        order_list.save(update_fields=["finalized_at", "finalized_by"])
        messages.success(request, "List unfinalized. Editing re-enabled for employees.")
    return redirect("inventory:weekly_list_detail", list_id=list_id)


# ===== Admin-only: basic employee management =====

@login_required
@user_passes_test(_staff_required)
def user_manage(request):
    """Admin UI to create users (employee or admin) and reset passwords."""
    User = get_user_model()
    context = {"active_tab": "users"}

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create":
            username = (request.POST.get("username") or "").strip()
            pw1 = request.POST.get("password1") or ""
            pw2 = request.POST.get("password2") or ""
            role = (request.POST.get("role") or "employee").strip().lower()
            if not username or not pw1:
                context["error"] = "Username and password are required."
            elif pw1 != pw2:
                context["error"] = "Passwords do not match."
            elif User.objects.filter(username=username).exists():
                context["error"] = "Username already exists."
            else:
                user = User.objects.create_user(username=username, password=pw1)
                is_admin = role == "admin"
                # Grant staff for admins; do not automatically grant superuser.
                user.is_staff = bool(is_admin)
                user.save(update_fields=["is_staff"])
                role_label = "Admin" if is_admin else "Employee"
                context["success"] = f"{role_label} '{username}' created."
        elif action == "reset":
            try:
                target_id = int(request.POST.get("user_id") or "0")
            except ValueError:
                target_id = 0
            pw1 = request.POST.get("new_password1") or ""
            pw2 = request.POST.get("new_password2") or ""
            if not target_id or not pw1:
                context["error"] = "User and new password are required."
            elif pw1 != pw2:
                context["error"] = "Passwords do not match."
            else:
                try:
                    target = User.objects.get(pk=target_id)
                    target.set_password(pw1)
                    target.save(update_fields=["password"])
                    context["success"] = f"Password reset for '{target.username}'."
                except User.DoesNotExist:
                    context["error"] = "User not found."

    # Prepare lists for display: admins and employees
    employees = get_user_model().objects.filter(is_staff=False, is_active=True).order_by("username")
    admins = get_user_model().objects.filter(is_staff=True, is_active=True).order_by("username")
    context.update({"employees": employees, "admins": admins})

    return render(request, "inventory/users_manage.html", context)


@login_required
def logout_view(request):
    """Log out the current user; requires POST to mitigate CSRF.

    Example:
        curl -X POST http://localhost:8000/accounts/logout/
    """
    from django.shortcuts import redirect
    from django.views.decorators.http import require_POST

    # Only allow POST to prevent CSRF attacks via GET links/images
    if request.method == "POST":
        logout(request)
        return redirect("inventory:home")
    return JsonResponse({"error": "Method not allowed. Use POST."}, status=405)
